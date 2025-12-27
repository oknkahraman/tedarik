from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import asyncio
import shutil
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from io import BytesIO
import resend
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from data import (
    MATERIALS, FORM_TYPES, MANUFACTURING_METHODS,
    PROJECT_STATUSES, PART_STATUSES, ORDER_STATUSES, QUOTE_STATUSES, CURRENCIES
)

ROOT_DIR = Path(__file__).parent
UPLOADS_DIR = ROOT_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Resend setup
resend_api_key = os.environ.get('RESEND_API_KEY')
if resend_api_key:
    resend.api_key = resend_api_key
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')

app = FastAPI(title="ProManufakt API", version="1.0.0")
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ===================== PYDANTIC MODELS =====================

# Project Models
class ProjectCreate(BaseModel):
    name: str
    customer_name: Optional[str] = None
    start_date: str
    end_date: str
    notes: Optional[str] = None

class Project(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    code: str = ""
    name: str
    customer_name: Optional[str] = None
    start_date: str
    end_date: str
    status: str = "planning"
    notes: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Part Models
class PartDimensions(BaseModel):
    width: Optional[float] = None
    height: Optional[float] = None
    length: Optional[float] = None
    diameter: Optional[float] = None
    outer_diameter: Optional[float] = None
    inner_diameter: Optional[float] = None

class PartCreate(BaseModel):
    project_id: str
    name: str
    code: str
    quantity: int
    material: Optional[str] = None
    form_type: Optional[str] = None
    dimensions: Optional[PartDimensions] = None
    manufacturing_methods: List[str] = []
    notes: Optional[str] = None
    technical_drawing_url: Optional[str] = None

class Part(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project_id: str
    name: str
    code: str
    quantity: int
    material: Optional[str] = None
    form_type: Optional[str] = None
    dimensions: Optional[PartDimensions] = None
    manufacturing_methods: List[str] = []
    status: str = "pending"
    notes: Optional[str] = None
    technical_drawing_url: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Supplier Models
class SupplierCreate(BaseModel):
    name: str
    contact_person: str
    email: EmailStr
    phone: Optional[str] = None
    address: Optional[str] = None
    tax_id: Optional[str] = None
    specializations: List[str] = []
    payment_terms: Optional[int] = 30
    notes: Optional[str] = None

class SupplierPerformance(BaseModel):
    total_orders: int = 0
    on_time_deliveries: int = 0
    quality_rejections: int = 0
    average_price_ratio: float = 1.0
    delivery_score: float = 40.0
    quality_score: float = 30.0
    price_score: float = 15.0
    payment_score: float = 10.0
    total_score: float = 95.0

class Supplier(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    contact_person: str
    email: EmailStr
    phone: Optional[str] = None
    address: Optional[str] = None
    tax_id: Optional[str] = None
    specializations: List[str] = []
    payment_terms: int = 30
    performance: SupplierPerformance = Field(default_factory=SupplierPerformance)
    notes: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Quote Models
class QuoteRequestCreate(BaseModel):
    part_id: str
    supplier_ids: List[str]
    manufacturing_method: str
    deadline: str
    notes: Optional[str] = None

class QuoteResponseCreate(BaseModel):
    quote_request_id: str
    supplier_id: str
    unit_price: float
    currency: str = "TRY"
    total_price: float
    delivery_date: str
    payment_terms: int = 30
    notes: Optional[str] = None

class QuoteRequest(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    part_id: str
    supplier_ids: List[str]
    manufacturing_method: str
    deadline: str
    status: str = "requested"
    notes: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class QuoteResponse(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    quote_request_id: str
    supplier_id: str
    unit_price: float
    currency: str = "TRY"
    total_price: float
    delivery_date: str
    payment_terms: int = 30
    status: str = "received"
    notes: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Order Models
class OrderCreate(BaseModel):
    quote_response_id: str
    part_id: str
    supplier_id: str
    quantity: int
    unit_price: float
    currency: str = "TRY"
    total_price: float
    expected_delivery: str
    notes: Optional[str] = None

class Order(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    code: str = ""
    quote_response_id: str
    part_id: str
    supplier_id: str
    quantity: int
    unit_price: float
    currency: str = "TRY"
    total_price: float
    expected_delivery: str
    actual_delivery: Optional[str] = None
    status: str = "pending"
    notes: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Notification Models
class Notification(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type: str
    title: str
    message: str
    reference_type: Optional[str] = None
    reference_id: Optional[str] = None
    is_read: bool = False
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Currency Rate Models
class CurrencyRateUpdate(BaseModel):
    usd_to_try: float
    eur_to_try: float

class CurrencyRate(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    usd_to_try: float
    eur_to_try: float
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_by: str = "manual"

# Settings Models
class SettingsUpdate(BaseModel):
    company_name: Optional[str] = None
    company_email: Optional[str] = None
    company_phone: Optional[str] = None
    company_address: Optional[str] = None

class Settings(BaseModel):
    id: str = "settings"
    company_name: str = "ProManufakt"
    company_email: str = ""
    company_phone: str = ""
    company_address: str = ""

# Email Models
class EmailRequest(BaseModel):
    recipient_email: EmailStr
    subject: str
    html_content: str

# ===================== HELPER FUNCTIONS =====================

async def generate_project_code():
    year = datetime.now().year
    count = await db.projects.count_documents({"code": {"$regex": f"^PRJ-{year}"}})
    return f"PRJ-{year}-{str(count + 1).zfill(3)}"

async def generate_order_code():
    year = datetime.now().year
    count = await db.orders.count_documents({"code": {"$regex": f"^SIP-{year}"}})
    return f"SIP-{year}-{str(count + 1).zfill(3)}"

async def create_notification(notification_type: str, title: str, message: str, ref_type: str = None, ref_id: str = None):
    notification = Notification(
        type=notification_type,
        title=title,
        message=message,
        reference_type=ref_type,
        reference_id=ref_id
    )
    await db.notifications.insert_one(notification.model_dump())
    return notification

def calculate_supplier_score(performance: dict) -> dict:
    total_orders = performance.get("total_orders", 0)
    on_time = performance.get("on_time_deliveries", 0)
    rejections = performance.get("quality_rejections", 0)
    price_ratio = performance.get("average_price_ratio", 1.0)
    payment_terms = performance.get("payment_terms", 30)
    
    # Delivery Score (40 points)
    if total_orders > 0:
        delivery_rate = on_time / total_orders
        delivery_score = delivery_rate * 40
    else:
        delivery_score = 40.0
    
    # Quality Score (30 points)
    if total_orders > 0:
        rejection_rate = rejections / total_orders
        quality_score = (1 - rejection_rate) * 30
    else:
        quality_score = 30.0
    
    # Price Score (20 points)
    if price_ratio <= 0.95:
        price_score = 20.0
    elif price_ratio <= 1.0:
        price_score = 15.0
    elif price_ratio <= 1.1:
        price_score = 10.0
    else:
        price_score = 5.0
    
    # Payment Terms Score (10 points)
    if payment_terms >= 90:
        payment_score = 10.0
    elif payment_terms >= 60:
        payment_score = 8.0
    elif payment_terms >= 30:
        payment_score = 5.0
    else:
        payment_score = 3.0
    
    total_score = delivery_score + quality_score + price_score + payment_score
    
    return {
        "delivery_score": round(delivery_score, 1),
        "quality_score": round(quality_score, 1),
        "price_score": round(price_score, 1),
        "payment_score": round(payment_score, 1),
        "total_score": round(total_score, 1)
    }

# ===================== API ROUTES =====================

@api_router.get("/")
async def root():
    return {"message": "ProManufakt API", "version": "1.0.0"}

# --- Static Data Routes ---
@api_router.get("/materials")
async def get_materials():
    return MATERIALS

@api_router.get("/form-types")
async def get_form_types():
    return FORM_TYPES

@api_router.get("/manufacturing-methods")
async def get_manufacturing_methods():
    return MANUFACTURING_METHODS

@api_router.get("/statuses")
async def get_statuses():
    return {
        "project": PROJECT_STATUSES,
        "part": PART_STATUSES,
        "order": ORDER_STATUSES,
        "quote": QUOTE_STATUSES
    }

@api_router.get("/currencies")
async def get_currencies():
    return CURRENCIES

# --- Project Routes ---
@api_router.post("/projects", response_model=Project)
async def create_project(data: ProjectCreate):
    code = await generate_project_code()
    project = Project(code=code, **data.model_dump())
    doc = project.model_dump()
    await db.projects.insert_one(doc)
    await create_notification("project", "Yeni Proje", f"{project.name} projesi oluşturuldu", "project", project.id)
    return project

@api_router.get("/projects")
async def get_projects(status: Optional[str] = None):
    query = {}
    if status:
        query["status"] = status
    projects = await db.projects.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return projects

@api_router.get("/projects/{project_id}")
async def get_project(project_id: str):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    return project

@api_router.put("/projects/{project_id}")
async def update_project(project_id: str, data: dict):
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.projects.update_one({"id": project_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    return await db.projects.find_one({"id": project_id}, {"_id": 0})

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str):
    result = await db.projects.delete_one({"id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    await db.parts.delete_many({"project_id": project_id})
    return {"message": "Proje silindi"}

# --- Part Routes ---
@api_router.post("/parts", response_model=Part)
async def create_part(data: PartCreate):
    # Check if project exists
    project = await db.projects.find_one({"id": data.project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    part = Part(**data.model_dump())
    doc = part.model_dump()
    if doc.get("dimensions"):
        doc["dimensions"] = doc["dimensions"]
    await db.parts.insert_one(doc)
    return part

@api_router.get("/parts")
async def get_parts(project_id: Optional[str] = None, status: Optional[str] = None):
    query = {}
    if project_id:
        query["project_id"] = project_id
    if status:
        query["status"] = status
    parts = await db.parts.find(query, {"_id": 0}).to_list(1000)
    return parts

@api_router.get("/parts/{part_id}")
async def get_part(part_id: str):
    part = await db.parts.find_one({"id": part_id}, {"_id": 0})
    if not part:
        raise HTTPException(status_code=404, detail="Parça bulunamadı")
    return part

@api_router.put("/parts/{part_id}")
async def update_part(part_id: str, data: dict):
    result = await db.parts.update_one({"id": part_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Parça bulunamadı")
    return await db.parts.find_one({"id": part_id}, {"_id": 0})

@api_router.delete("/parts/{part_id}")
async def delete_part(part_id: str):
    # Get part to delete associated files
    part = await db.parts.find_one({"id": part_id}, {"_id": 0})
    if part:
        # Delete technical drawing file if exists
        if part.get("technical_drawing_filename"):
            file_path = UPLOADS_DIR / part["technical_drawing_filename"]
            if file_path.exists():
                file_path.unlink()
        # Delete additional documents
        for doc_file in part.get("additional_documents", []):
            if doc_file.get("filename"):
                file_path = UPLOADS_DIR / doc_file["filename"]
                if file_path.exists():
                    file_path.unlink()
    
    result = await db.parts.delete_one({"id": part_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Parça bulunamadı")
    return {"message": "Parça silindi"}

# --- File Upload Routes ---
ALLOWED_EXTENSIONS = {'.pdf', '.dwg', '.dxf', '.step', '.stp', '.iges', '.igs', '.png', '.jpg', '.jpeg'}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB

@api_router.post("/upload/technical-drawing/{part_id}")
async def upload_technical_drawing(part_id: str, file: UploadFile = File(...)):
    """Upload technical drawing for a part (PDF, DWG, DXF, STEP, etc.)"""
    part = await db.parts.find_one({"id": part_id}, {"_id": 0})
    if not part:
        raise HTTPException(status_code=404, detail="Parça bulunamadı")
    
    # Validate file extension
    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400, 
            detail=f"Desteklenmeyen dosya formatı. İzin verilen: {', '.join(ALLOWED_EXTENSIONS)}"
        )
    
    # Generate unique filename
    unique_filename = f"{part_id}_drawing_{uuid.uuid4().hex[:8]}{file_ext}"
    file_path = UPLOADS_DIR / unique_filename
    
    # Delete old file if exists
    if part.get("technical_drawing_filename"):
        old_file = UPLOADS_DIR / part["technical_drawing_filename"]
        if old_file.exists():
            old_file.unlink()
    
    # Save file
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
    except Exception as e:
        logger.error(f"File upload error: {str(e)}")
        raise HTTPException(status_code=500, detail="Dosya yüklenemedi")
    
    # Update part with file info
    await db.parts.update_one(
        {"id": part_id},
        {"$set": {
            "technical_drawing_filename": unique_filename,
            "technical_drawing_original_name": file.filename,
            "technical_drawing_uploaded_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "success": True,
        "filename": unique_filename,
        "original_name": file.filename,
        "message": "Teknik resim başarıyla yüklendi"
    }

@api_router.post("/upload/document/{part_id}")
async def upload_additional_document(part_id: str, file: UploadFile = File(...)):
    """Upload additional document for a part"""
    part = await db.parts.find_one({"id": part_id}, {"_id": 0})
    if not part:
        raise HTTPException(status_code=404, detail="Parça bulunamadı")
    
    # Validate file extension
    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400, 
            detail=f"Desteklenmeyen dosya formatı. İzin verilen: {', '.join(ALLOWED_EXTENSIONS)}"
        )
    
    # Generate unique filename
    unique_filename = f"{part_id}_doc_{uuid.uuid4().hex[:8]}{file_ext}"
    file_path = UPLOADS_DIR / unique_filename
    
    # Save file
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
    except Exception as e:
        logger.error(f"File upload error: {str(e)}")
        raise HTTPException(status_code=500, detail="Dosya yüklenemedi")
    
    # Add to additional documents list
    doc_info = {
        "id": str(uuid.uuid4()),
        "filename": unique_filename,
        "original_name": file.filename,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.parts.update_one(
        {"id": part_id},
        {"$push": {"additional_documents": doc_info}}
    )
    
    return {
        "success": True,
        "document": doc_info,
        "message": "Döküman başarıyla yüklendi"
    }

@api_router.get("/files/{filename}")
async def get_file(filename: str):
    """Download a file"""
    file_path = UPLOADS_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Dosya bulunamadı")
    
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/octet-stream"
    )

@api_router.delete("/files/{part_id}/{filename}")
async def delete_file(part_id: str, filename: str):
    """Delete a file from a part"""
    part = await db.parts.find_one({"id": part_id}, {"_id": 0})
    if not part:
        raise HTTPException(status_code=404, detail="Parça bulunamadı")
    
    file_path = UPLOADS_DIR / filename
    
    # Check if it's the technical drawing
    if part.get("technical_drawing_filename") == filename:
        if file_path.exists():
            file_path.unlink()
        await db.parts.update_one(
            {"id": part_id},
            {"$unset": {
                "technical_drawing_filename": "",
                "technical_drawing_original_name": "",
                "technical_drawing_uploaded_at": ""
            }}
        )
        return {"success": True, "message": "Teknik resim silindi"}
    
    # Check if it's an additional document
    docs = part.get("additional_documents", [])
    doc_to_remove = None
    for doc in docs:
        if doc.get("filename") == filename:
            doc_to_remove = doc
            break
    
    if doc_to_remove:
        if file_path.exists():
            file_path.unlink()
        await db.parts.update_one(
            {"id": part_id},
            {"$pull": {"additional_documents": {"filename": filename}}}
        )
        return {"success": True, "message": "Döküman silindi"}
    
    raise HTTPException(status_code=404, detail="Dosya bu parçaya ait değil")

# --- Supplier Routes ---
@api_router.post("/suppliers", response_model=Supplier)
async def create_supplier(data: SupplierCreate):
    supplier = Supplier(**data.model_dump())
    doc = supplier.model_dump()
    await db.suppliers.insert_one(doc)
    return supplier

@api_router.get("/suppliers")
async def get_suppliers(specialization: Optional[str] = None):
    query = {}
    if specialization:
        query["specializations"] = specialization
    suppliers = await db.suppliers.find(query, {"_id": 0}).to_list(1000)
    return suppliers

@api_router.get("/suppliers/{supplier_id}")
async def get_supplier(supplier_id: str):
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Tedarikçi bulunamadı")
    return supplier

@api_router.put("/suppliers/{supplier_id}")
async def update_supplier(supplier_id: str, data: dict):
    result = await db.suppliers.update_one({"id": supplier_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tedarikçi bulunamadı")
    return await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str):
    result = await db.suppliers.delete_one({"id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tedarikçi bulunamadı")
    return {"message": "Tedarikçi silindi"}

# --- Quote Routes ---
@api_router.post("/quote-requests", response_model=QuoteRequest)
async def create_quote_request(data: QuoteRequestCreate):
    quote_request = QuoteRequest(**data.model_dump())
    doc = quote_request.model_dump()
    await db.quote_requests.insert_one(doc)
    
    # Create notifications for each supplier
    part = await db.parts.find_one({"id": data.part_id}, {"_id": 0})
    for supplier_id in data.supplier_ids:
        supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
        if supplier and part:
            await create_notification(
                "quote_request",
                "Teklif Talebi",
                f"{part.get('name', '')} parçası için teklif talebi oluşturuldu",
                "quote_request",
                quote_request.id
            )
    
    return quote_request

@api_router.get("/quote-requests")
async def get_quote_requests(part_id: Optional[str] = None, status: Optional[str] = None):
    query = {}
    if part_id:
        query["part_id"] = part_id
    if status:
        query["status"] = status
    requests = await db.quote_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return requests

@api_router.post("/quote-responses", response_model=QuoteResponse)
async def create_quote_response(data: QuoteResponseCreate):
    quote_response = QuoteResponse(**data.model_dump())
    doc = quote_response.model_dump()
    await db.quote_responses.insert_one(doc)
    
    # Update quote request status
    await db.quote_requests.update_one(
        {"id": data.quote_request_id},
        {"$set": {"status": "received"}}
    )
    
    return quote_response

@api_router.get("/quote-responses")
async def get_quote_responses(quote_request_id: Optional[str] = None):
    query = {}
    if quote_request_id:
        query["quote_request_id"] = quote_request_id
    responses = await db.quote_responses.find(query, {"_id": 0}).to_list(1000)
    return responses

@api_router.get("/quote-comparison/{quote_request_id}")
async def get_quote_comparison(quote_request_id: str):
    """Compare quotes for a specific request with scoring"""
    quote_request = await db.quote_requests.find_one({"id": quote_request_id}, {"_id": 0})
    if not quote_request:
        raise HTTPException(status_code=404, detail="Teklif talebi bulunamadı")
    
    responses = await db.quote_responses.find({"quote_request_id": quote_request_id}, {"_id": 0}).to_list(100)
    
    # Get currency rates
    rates = await db.currency_rates.find_one({}, {"_id": 0}, sort=[("updated_at", -1)])
    usd_rate = rates.get("usd_to_try", 33.0) if rates else 33.0
    eur_rate = rates.get("eur_to_try", 36.5) if rates else 36.5
    
    # Calculate comparison data
    comparison = []
    min_price_try = float('inf')
    
    for resp in responses:
        supplier = await db.suppliers.find_one({"id": resp["supplier_id"]}, {"_id": 0})
        
        # Convert to TRY
        if resp["currency"] == "USD":
            price_try = resp["total_price"] * usd_rate
        elif resp["currency"] == "EUR":
            price_try = resp["total_price"] * eur_rate
        else:
            price_try = resp["total_price"]
        
        if price_try < min_price_try:
            min_price_try = price_try
        
        comparison.append({
            "response": resp,
            "supplier": supplier,
            "price_try": price_try
        })
    
    # Calculate scores
    deadline = datetime.fromisoformat(quote_request["deadline"].replace("Z", "+00:00"))
    
    for item in comparison:
        resp = item["response"]
        supplier = item["supplier"]
        
        # Price score (40%)
        if min_price_try > 0:
            price_ratio = min_price_try / item["price_try"]
            price_score = price_ratio * 40
        else:
            price_score = 40
        
        # Delivery score (30%)
        delivery_date = datetime.fromisoformat(resp["delivery_date"].replace("Z", "+00:00"))
        days_diff = (delivery_date - deadline).days
        if days_diff <= 0:
            delivery_score = 30 + min(abs(days_diff) * 2, 6)  # Bonus for early
        else:
            delivery_score = max(30 - days_diff * 3, 0)
        
        # Quality score (20%) - from supplier performance
        if supplier and supplier.get("performance"):
            quality_score = supplier["performance"].get("quality_score", 24) * 20 / 30
        else:
            quality_score = 16
        
        # Payment score (10%)
        payment_terms = resp.get("payment_terms", 30)
        if payment_terms >= 90:
            payment_score = 10
        elif payment_terms >= 60:
            payment_score = 8
        elif payment_terms >= 30:
            payment_score = 5
        else:
            payment_score = 3
        
        item["scores"] = {
            "price": round(price_score, 1),
            "delivery": round(delivery_score, 1),
            "quality": round(quality_score, 1),
            "payment": round(payment_score, 1),
            "total": round(price_score + delivery_score + quality_score + payment_score, 1)
        }
    
    # Sort by total score
    comparison.sort(key=lambda x: x["scores"]["total"], reverse=True)
    
    return {
        "quote_request": quote_request,
        "comparison": comparison,
        "currency_rates": {"usd": usd_rate, "eur": eur_rate}
    }

# --- Order Routes ---
@api_router.post("/orders", response_model=Order)
async def create_order(data: OrderCreate):
    code = await generate_order_code()
    order = Order(code=code, **data.model_dump())
    doc = order.model_dump()
    await db.orders.insert_one(doc)
    
    # Update quote response status
    await db.quote_responses.update_one(
        {"id": data.quote_response_id},
        {"$set": {"status": "approved"}}
    )
    
    # Update part status
    await db.parts.update_one(
        {"id": data.part_id},
        {"$set": {"status": "in_production"}}
    )
    
    # Create notification
    supplier = await db.suppliers.find_one({"id": data.supplier_id}, {"_id": 0})
    part = await db.parts.find_one({"id": data.part_id}, {"_id": 0})
    await create_notification(
        "order",
        "Sipariş Oluşturuldu",
        f"{code}: {part.get('name', '')} siparişi {supplier.get('name', '')} tedarikçisine verildi",
        "order",
        order.id
    )
    
    return order

@api_router.get("/orders")
async def get_orders(status: Optional[str] = None, supplier_id: Optional[str] = None):
    query = {}
    if status:
        query["status"] = status
    if supplier_id:
        query["supplier_id"] = supplier_id
    orders = await db.orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return orders

@api_router.get("/orders/{order_id}")
async def get_order(order_id: str):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
    return order

@api_router.put("/orders/{order_id}")
async def update_order(order_id: str, data: dict):
    result = await db.orders.update_one({"id": order_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
    
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    
    # If order is delivered, update supplier performance
    if data.get("status") == "delivered":
        supplier_id = order["supplier_id"]
        supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
        if supplier:
            perf = supplier.get("performance", {})
            perf["total_orders"] = perf.get("total_orders", 0) + 1
            
            # Check if on time
            if order.get("actual_delivery") and order.get("expected_delivery"):
                actual = datetime.fromisoformat(order["actual_delivery"].replace("Z", "+00:00"))
                expected = datetime.fromisoformat(order["expected_delivery"].replace("Z", "+00:00"))
                if actual <= expected:
                    perf["on_time_deliveries"] = perf.get("on_time_deliveries", 0) + 1
            
            # Recalculate scores
            scores = calculate_supplier_score(perf)
            perf.update(scores)
            
            await db.suppliers.update_one(
                {"id": supplier_id},
                {"$set": {"performance": perf}}
            )
        
        # Update part status
        await db.parts.update_one(
            {"id": order["part_id"]},
            {"$set": {"status": "completed"}}
        )
    
    return order

@api_router.delete("/orders/{order_id}")
async def delete_order(order_id: str):
    result = await db.orders.delete_one({"id": order_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
    return {"message": "Sipariş silindi"}

# --- Notification Routes ---
@api_router.get("/notifications")
async def get_notifications(is_read: Optional[bool] = None):
    query = {}
    if is_read is not None:
        query["is_read"] = is_read
    notifications = await db.notifications.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return notifications

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str):
    await db.notifications.update_one({"id": notification_id}, {"$set": {"is_read": True}})
    return {"message": "Bildirim okundu olarak işaretlendi"}

@api_router.put("/notifications/read-all")
async def mark_all_notifications_read():
    await db.notifications.update_many({}, {"$set": {"is_read": True}})
    return {"message": "Tüm bildirimler okundu olarak işaretlendi"}

# --- Currency Rate Routes ---
@api_router.get("/currency-rates")
async def get_currency_rates():
    rates = await db.currency_rates.find_one({}, {"_id": 0}, sort=[("updated_at", -1)])
    if not rates:
        return {"usd_to_try": 33.0, "eur_to_try": 36.5, "updated_at": datetime.now(timezone.utc).isoformat()}
    return rates

@api_router.post("/currency-rates")
async def update_currency_rates(data: CurrencyRateUpdate):
    try:
        rate = CurrencyRate(
            usd_to_try=data.usd_to_try,
            eur_to_try=data.eur_to_try
        )
        doc = rate.model_dump()
        await db.currency_rates.insert_one(doc)
        # Return without _id
        return {
            "id": doc["id"],
            "usd_to_try": doc["usd_to_try"],
            "eur_to_try": doc["eur_to_try"],
            "updated_at": doc["updated_at"],
            "updated_by": doc["updated_by"]
        }
    except Exception as e:
        logger.error(f"Currency rate update error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Döviz kuru güncellenemedi: {str(e)}")

# --- Settings Routes ---
@api_router.get("/settings")
async def get_settings():
    settings = await db.settings.find_one({"id": "settings"}, {"_id": 0})
    if not settings:
        return Settings().model_dump()
    return settings

@api_router.put("/settings")
async def update_settings(data: SettingsUpdate):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    await db.settings.update_one(
        {"id": "settings"},
        {"$set": update_data},
        upsert=True
    )
    return await db.settings.find_one({"id": "settings"}, {"_id": 0})

# --- Email Routes ---
@api_router.post("/send-email")
async def send_email(request: EmailRequest):
    if not resend_api_key:
        raise HTTPException(status_code=500, detail="E-posta servisi yapılandırılmamış")
    
    params = {
        "from": SENDER_EMAIL,
        "to": [request.recipient_email],
        "subject": request.subject,
        "html": request.html_content
    }
    
    try:
        email = await asyncio.to_thread(resend.Emails.send, params)
        return {
            "status": "success",
            "message": f"E-posta {request.recipient_email} adresine gönderildi",
            "email_id": email.get("id")
        }
    except Exception as e:
        logger.error(f"E-posta gönderme hatası: {str(e)}")
        raise HTTPException(status_code=500, detail=f"E-posta gönderilemedi: {str(e)}")

# --- Dashboard Routes ---
@api_router.get("/dashboard/stats")
async def get_dashboard_stats():
    # Project stats
    total_projects = await db.projects.count_documents({})
    active_projects = await db.projects.count_documents({"status": {"$in": ["planning", "in_progress"]}})
    
    # Part stats
    total_parts = await db.parts.count_documents({})
    parts_in_production = await db.parts.count_documents({"status": "in_production"})
    
    # Order stats
    total_orders = await db.orders.count_documents({})
    pending_orders = await db.orders.count_documents({"status": {"$in": ["pending", "confirmed", "in_production"]}})
    
    # Supplier stats
    total_suppliers = await db.suppliers.count_documents({})
    
    # Upcoming deadlines (next 7 days)
    now = datetime.now(timezone.utc)
    next_week = now + timedelta(days=7)
    
    upcoming_orders = await db.orders.find({
        "status": {"$in": ["pending", "confirmed", "in_production"]},
        "expected_delivery": {
            "$gte": now.isoformat(),
            "$lte": next_week.isoformat()
        }
    }, {"_id": 0}).to_list(10)
    
    # Pending quotes
    pending_quotes = await db.quote_requests.count_documents({"status": "requested"})
    
    # Unread notifications
    unread_notifications = await db.notifications.count_documents({"is_read": False})
    
    return {
        "projects": {"total": total_projects, "active": active_projects},
        "parts": {"total": total_parts, "in_production": parts_in_production},
        "orders": {"total": total_orders, "pending": pending_orders},
        "suppliers": {"total": total_suppliers},
        "pending_quotes": pending_quotes,
        "upcoming_orders": upcoming_orders,
        "unread_notifications": unread_notifications
    }

@api_router.get("/dashboard/gantt/{project_id}")
async def get_gantt_data(project_id: str):
    """Get Gantt chart data for a project"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    parts = await db.parts.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    gantt_data = []
    current_date = datetime.fromisoformat(project["start_date"].replace("Z", "+00:00"))
    
    for part in parts:
        part_start = current_date
        methods = part.get("manufacturing_methods", [])
        
        for method_code in methods:
            method = MANUFACTURING_METHODS.get(method_code)
            if method:
                duration = method.get("duration_days", 2)
                end_date = part_start + timedelta(days=duration)
                
                gantt_data.append({
                    "id": f"{part['id']}_{method_code}",
                    "part_id": part["id"],
                    "part_name": part["name"],
                    "part_code": part["code"],
                    "method_code": method_code,
                    "method_name": method["name"],
                    "category": method["category"],
                    "start": part_start.isoformat(),
                    "end": end_date.isoformat(),
                    "duration_days": duration
                })
                
                part_start = end_date
    
    return {
        "project": project,
        "tasks": gantt_data
    }

# --- Excel Import/Export Routes ---
@api_router.get("/export/parts/{project_id}")
async def export_parts_to_excel(project_id: str):
    """Export project parts to Excel file"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    parts = await db.parts.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parça Listesi"
    
    # Header style
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Parça Kodu", "Parça Adı", "Adet", "Malzeme", "Form Tipi", "Ölçü 1", "Ölçü 2", "Ölçü 3", "İmalat 1", "İmalat 2", "İmalat 3", "İmalat 4", "İmalat 5", "Notlar"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data rows
    for row_idx, part in enumerate(parts, 2):
        dims = part.get("dimensions", {}) or {}
        methods = part.get("manufacturing_methods", [])
        
        # Get dimension values based on form type
        dim_values = []
        form_type = part.get("form_type", "")
        if form_type == "prizmatik":
            dim_values = [dims.get("width"), dims.get("height"), dims.get("length")]
        elif form_type == "silindirik":
            dim_values = [dims.get("diameter"), dims.get("length"), None]
        elif form_type == "boru":
            dim_values = [dims.get("outer_diameter"), dims.get("inner_diameter"), dims.get("length")]
        else:
            dim_values = [None, None, None]
        
        row_data = [
            part.get("code", ""),
            part.get("name", ""),
            part.get("quantity", 0),
            part.get("material", ""),
            form_type,
            dim_values[0] if len(dim_values) > 0 else None,
            dim_values[1] if len(dim_values) > 1 else None,
            dim_values[2] if len(dim_values) > 2 else None,
            methods[0] if len(methods) > 0 else None,
            methods[1] if len(methods) > 1 else None,
            methods[2] if len(methods) > 2 else None,
            methods[3] if len(methods) > 3 else None,
            methods[4] if len(methods) > 4 else None,
            part.get("notes", "")
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{project['code']}_parcalar.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/export/template")
async def export_parts_template():
    """Export empty template for parts import"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parça Şablonu"
    
    # Header style
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers
    headers = ["Parça Kodu*", "Parça Adı*", "Adet*", "Malzeme", "Form Tipi", "Ölçü 1", "Ölçü 2", "Ölçü 3", "İmalat 1", "İmalat 2", "İmalat 3", "İmalat 4", "İmalat 5", "Notlar"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Instructions sheet
    ws2 = wb.create_sheet("Açıklamalar")
    instructions = [
        ["ALAN", "AÇIKLAMA", "ÖRNEK"],
        ["Parça Kodu*", "Benzersiz parça kodu (zorunlu)", "MIL-001"],
        ["Parça Adı*", "Parça adı (zorunlu)", "Mil"],
        ["Adet*", "Adet sayısı (zorunlu)", "50"],
        ["Malzeme", "Malzeme kodu", "AISI304, S355, AL5754"],
        ["Form Tipi", "prizmatik / silindirik / boru", "silindirik"],
        ["Ölçü 1", "Prizmatik: Genişlik, Silindirik: Çap, Boru: Dış Çap", "50"],
        ["Ölçü 2", "Prizmatik: Yükseklik, Silindirik: Uzunluk, Boru: İç Çap", "200"],
        ["Ölçü 3", "Prizmatik: Uzunluk, Boru: Uzunluk", "300"],
        ["İmalat 1-5", "İmalat yöntemi kodları (sıralı)", "3001, 5009, 5002"],
        ["", "", ""],
        ["İMALAT KODLARI", "", ""],
        ["1001", "Lazer Kesim", ""],
        ["1002", "Abkant Büküm", ""],
        ["2001", "MIG/MAG Kaynak", ""],
        ["2002", "TIG Kaynak", ""],
        ["3001", "CNC Torna", ""],
        ["3002", "CNC Dik İşlem (3 Eksen)", ""],
        ["5001", "Eloksal", ""],
        ["5002", "Galvaniz", ""],
        ["5003", "Statik Boya", ""],
        ["5009", "Isıl İşlem", ""],
        ["9001", "Rulman", ""],
        ["9007", "Çelik Ham Malzeme", ""],
    ]
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)
    
    # Adjust column widths
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=parca_sablonu.xlsx"}
    )

@api_router.post("/import/parts/{project_id}")
async def import_parts_from_excel(project_id: str, file: UploadFile = File(...)):
    """Import parts from Excel file"""
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Proje bulunamadı")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Sadece Excel dosyası (.xlsx, .xls) kabul edilir")
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        imported = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[0] or not row[1] or not row[2]:
                continue  # Skip empty rows
            
            try:
                code = str(row[0]).strip()
                name = str(row[1]).strip()
                quantity = int(row[2]) if row[2] else 1
                material = str(row[3]).strip() if row[3] else None
                form_type = str(row[4]).strip().lower() if row[4] else None
                
                # Parse dimensions based on form type
                dimensions = {}
                if form_type == "prizmatik":
                    dimensions = {
                        "width": float(row[5]) if row[5] else None,
                        "height": float(row[6]) if row[6] else None,
                        "length": float(row[7]) if row[7] else None
                    }
                elif form_type == "silindirik":
                    dimensions = {
                        "diameter": float(row[5]) if row[5] else None,
                        "length": float(row[6]) if row[6] else None
                    }
                elif form_type == "boru":
                    dimensions = {
                        "outer_diameter": float(row[5]) if row[5] else None,
                        "inner_diameter": float(row[6]) if row[6] else None,
                        "length": float(row[7]) if row[7] else None
                    }
                
                # Parse manufacturing methods
                methods = []
                for i in range(8, 13):
                    if row[i]:
                        methods.append(str(row[i]).strip())
                
                notes = str(row[13]).strip() if len(row) > 13 and row[13] else None
                
                # Check for duplicate code
                existing = await db.parts.find_one({"project_id": project_id, "code": code})
                if existing:
                    errors.append(f"Satır {row_idx}: '{code}' kodu zaten mevcut")
                    continue
                
                # Create part
                part = {
                    "id": str(uuid.uuid4()),
                    "project_id": project_id,
                    "code": code,
                    "name": name,
                    "quantity": quantity,
                    "material": material,
                    "form_type": form_type,
                    "dimensions": dimensions,
                    "manufacturing_methods": methods,
                    "status": "pending",
                    "notes": notes,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.parts.insert_one(part)
                imported += 1
                
            except Exception as e:
                errors.append(f"Satır {row_idx}: {str(e)}")
        
        return {
            "success": True,
            "imported": imported,
            "errors": errors,
            "message": f"{imported} parça başarıyla eklendi" + (f", {len(errors)} hata" if errors else "")
        }
        
    except Exception as e:
        logger.error(f"Excel import error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Excel dosyası okunamadı: {str(e)}")

# --- Quote Email Routes ---
class QuoteEmailRequest(BaseModel):
    quote_request_id: str
    supplier_ids: List[str]
    
@api_router.post("/send-quote-emails")
async def send_quote_emails(request: QuoteEmailRequest):
    """Send quote request emails to selected suppliers with form link"""
    if not resend_api_key:
        raise HTTPException(status_code=500, detail="E-posta servisi yapılandırılmamış. RESEND_API_KEY ekleyin.")
    
    quote_request = await db.quote_requests.find_one({"id": request.quote_request_id}, {"_id": 0})
    if not quote_request:
        raise HTTPException(status_code=404, detail="Teklif talebi bulunamadı")
    
    part = await db.parts.find_one({"id": quote_request["part_id"]}, {"_id": 0})
    if not part:
        raise HTTPException(status_code=404, detail="Parça bulunamadı")
    
    project = await db.projects.find_one({"id": part["project_id"]}, {"_id": 0})
    settings = await db.settings.find_one({"id": "settings"}, {"_id": 0}) or {}
    
    method = MANUFACTURING_METHODS.get(quote_request.get("manufacturing_method", ""), {})
    material = MATERIALS.get(part.get("material", ""), {})
    
    # Get app URL from environment or use default
    app_url = os.environ.get('APP_URL', 'https://supply-chain-84.preview.emergentagent.com')
    
    sent_emails = []
    failed_emails = []
    
    for supplier_id in request.supplier_ids:
        supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
        if not supplier:
            failed_emails.append({"supplier_id": supplier_id, "error": "Tedarikçi bulunamadı"})
            continue
        
        # Generate unique form token
        form_token = str(uuid.uuid4())[:8].upper()
        
        # Store token in quote request for validation
        await db.quote_requests.update_one(
            {"id": request.quote_request_id},
            {"$push": {"form_tokens": {"supplier_id": supplier_id, "token": form_token, "created_at": datetime.now(timezone.utc).isoformat()}}}
        )
        
        # Form link
        form_link = f"{app_url}/quote-form/{request.quote_request_id}?supplier={supplier_id}&token={form_token}"
        
        # Build email HTML
        html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; line-height: 1.6; color: #333; }}
        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
        .header {{ background: #1F4E79; color: white; padding: 20px; text-align: center; }}
        .content {{ padding: 20px; background: #f9f9f9; }}
        .section {{ background: white; padding: 15px; margin: 10px 0; border-left: 4px solid #1F4E79; }}
        .label {{ color: #666; font-size: 12px; text-transform: uppercase; }}
        .value {{ font-weight: bold; font-size: 16px; }}
        .button {{ display: inline-block; background: #F97316; color: white; padding: 15px 30px; text-decoration: none; border-radius: 5px; font-weight: bold; margin: 20px 0; }}
        .footer {{ text-align: center; padding: 20px; color: #666; font-size: 12px; }}
        table {{ width: 100%; border-collapse: collapse; }}
        td {{ padding: 8px; border-bottom: 1px solid #eee; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>TEKLİF TALEBİ</h1>
            <p>{settings.get('company_name', 'ProManufakt')}</p>
        </div>
        
        <div class="content">
            <p>Sayın <strong>{supplier.get('contact_person', supplier['name'])}</strong>,</p>
            
            <p>Aşağıdaki işlem için fiyat teklifinizi talep ediyoruz:</p>
            
            <div class="section">
                <p class="label">PROJE</p>
                <p class="value">{project.get('name', '-') if project else '-'} ({project.get('code', '-') if project else '-'})</p>
            </div>
            
            <div class="section">
                <p class="label">PARÇA BİLGİLERİ</p>
                <table>
                    <tr><td>Parça Adı:</td><td><strong>{part.get('name', '-')}</strong></td></tr>
                    <tr><td>Parça Kodu:</td><td><strong>{part.get('code', '-')}</strong></td></tr>
                    <tr><td>Miktar:</td><td><strong>{part.get('quantity', 0)} adet</strong></td></tr>
                    <tr><td>Malzeme:</td><td><strong>{material.get('name', part.get('material', '-'))}</strong></td></tr>
                </table>
            </div>
            
            <div class="section">
                <p class="label">İSTENEN İŞLEM</p>
                <p class="value">{method.get('code', '')} - {method.get('name', quote_request.get('manufacturing_method', '-'))}</p>
                <p style="color: #666; font-size: 14px;">{method.get('description', '')}</p>
            </div>
            
            <div class="section">
                <p class="label">TERMİN</p>
                <p class="value">Teklif Son Tarih: {quote_request.get('deadline', '-')[:10]}</p>
            </div>
            
            <center>
                <a href="{form_link}" class="button">TEKLİF FORMUNU DOLDUR</a>
            </center>
            
            <p style="text-align: center; color: #666; font-size: 12px;">
                Form Şifresi: <strong>{form_token}</strong>
            </p>
            
            {f'<p style="color: #666;"><strong>Not:</strong> {quote_request.get("notes", "")}</p>' if quote_request.get('notes') else ''}
        </div>
        
        <div class="footer">
            <p>{settings.get('company_name', 'ProManufakt')}</p>
            <p>{settings.get('company_email', '')} | {settings.get('company_phone', '')}</p>
            <p>{settings.get('company_address', '')}</p>
        </div>
    </div>
</body>
</html>
"""
        
        try:
            email_params = {
                "from": SENDER_EMAIL,
                "to": [supplier["email"]],
                "subject": f"Teklif Talebi - {part.get('code', '')} - {project.get('name', '') if project else ''}",
                "html": html_content
            }
            
            result = await asyncio.to_thread(resend.Emails.send, email_params)
            sent_emails.append({
                "supplier_id": supplier_id,
                "supplier_name": supplier["name"],
                "email": supplier["email"],
                "email_id": result.get("id")
            })
            
            # Create notification
            await db.notifications.insert_one({
                "id": str(uuid.uuid4()),
                "type": "email_sent",
                "title": "Teklif E-postası Gönderildi",
                "message": f"{supplier['name']} tedarikçisine teklif talebi gönderildi",
                "reference_type": "quote_request",
                "reference_id": request.quote_request_id,
                "is_read": False,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            
        except Exception as e:
            logger.error(f"Email send error to {supplier['email']}: {str(e)}")
            failed_emails.append({
                "supplier_id": supplier_id,
                "supplier_name": supplier["name"],
                "error": str(e)
            })
    
    return {
        "success": len(sent_emails) > 0,
        "sent": sent_emails,
        "failed": failed_emails,
        "message": f"{len(sent_emails)} e-posta gönderildi" + (f", {len(failed_emails)} başarısız" if failed_emails else "")
    }

# --- Public Quote Form Routes ---
@api_router.get("/quote-form/{quote_request_id}")
async def get_quote_form_data(quote_request_id: str, supplier: str, token: str):
    """Get quote form data for supplier to fill"""
    quote_request = await db.quote_requests.find_one({"id": quote_request_id}, {"_id": 0})
    if not quote_request:
        raise HTTPException(status_code=404, detail="Teklif talebi bulunamadı")
    
    # Validate token
    tokens = quote_request.get("form_tokens", [])
    valid_token = any(t.get("supplier_id") == supplier and t.get("token") == token for t in tokens)
    if not valid_token:
        raise HTTPException(status_code=403, detail="Geçersiz form linki")
    
    part = await db.parts.find_one({"id": quote_request["part_id"]}, {"_id": 0})
    supplier_data = await db.suppliers.find_one({"id": supplier}, {"_id": 0})
    project = await db.projects.find_one({"id": part["project_id"]}, {"_id": 0}) if part else None
    
    method = MANUFACTURING_METHODS.get(quote_request.get("manufacturing_method", ""), {})
    material = MATERIALS.get(part.get("material", ""), {}) if part else {}
    
    return {
        "quote_request": quote_request,
        "part": part,
        "project": project,
        "supplier": supplier_data,
        "method": method,
        "material": material
    }

class PublicQuoteSubmit(BaseModel):
    quote_request_id: str
    supplier_id: str
    token: str
    unit_price: float
    currency: str = "TRY"
    delivery_date: str
    payment_terms: int = 30
    notes: Optional[str] = None

@api_router.post("/quote-form/submit")
async def submit_quote_form(data: PublicQuoteSubmit):
    """Submit quote response from supplier form"""
    quote_request = await db.quote_requests.find_one({"id": data.quote_request_id}, {"_id": 0})
    if not quote_request:
        raise HTTPException(status_code=404, detail="Teklif talebi bulunamadı")
    
    # Validate token
    tokens = quote_request.get("form_tokens", [])
    valid_token = any(t.get("supplier_id") == data.supplier_id and t.get("token") == data.token for t in tokens)
    if not valid_token:
        raise HTTPException(status_code=403, detail="Geçersiz form linki")
    
    # Check if already submitted
    existing = await db.quote_responses.find_one({
        "quote_request_id": data.quote_request_id,
        "supplier_id": data.supplier_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Bu tedarikçi zaten teklif vermiş")
    
    part = await db.parts.find_one({"id": quote_request["part_id"]}, {"_id": 0})
    quantity = part.get("quantity", 1) if part else 1
    
    quote_response = {
        "id": str(uuid.uuid4()),
        "quote_request_id": data.quote_request_id,
        "supplier_id": data.supplier_id,
        "unit_price": data.unit_price,
        "currency": data.currency,
        "total_price": data.unit_price * quantity,
        "delivery_date": data.delivery_date,
        "payment_terms": data.payment_terms,
        "status": "received",
        "notes": data.notes,
        "submitted_via": "form",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.quote_responses.insert_one(quote_response)
    
    # Update quote request status
    await db.quote_requests.update_one(
        {"id": data.quote_request_id},
        {"$set": {"status": "received"}}
    )
    
    # Create notification
    supplier = await db.suppliers.find_one({"id": data.supplier_id}, {"_id": 0})
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()),
        "type": "quote_received",
        "title": "Yeni Teklif Alındı",
        "message": f"{supplier.get('name', 'Tedarikçi')} teklif gönderdi - {data.unit_price} {data.currency}",
        "reference_type": "quote_response",
        "reference_id": quote_response["id"],
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "success": True,
        "message": "Teklifiniz başarıyla gönderildi. Teşekkür ederiz!",
        "quote_response_id": quote_response["id"]
    }

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
